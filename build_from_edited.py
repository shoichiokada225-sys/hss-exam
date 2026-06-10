# -*- coding: utf-8 -*-
"""修正版Excel(HSS問題確認_全71問_重複削除済.xlsx)を唯一のソースに data/questions.json を再生成。
列: No, 出典, カテゴリ, 問題文, 正答, 正答番号, 選択肢1-5, 解説
出典が 'Word' を含む行 = 本試験(必出, source=honshiken-*), それ以外 = セットA(source=setA-*)
"""
import json, os, sys, re, difflib
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

ROOT = r"C:\Users\so\hss-exam"
SRC  = r"C:\Users\so\Downloads\HSS問題確認_全71問_重複削除済.xlsx"
DATA = os.path.join(ROOT, "data", "questions.json")

def norm(s): return re.sub(r"\s+", "", (s or "").strip())

# backup current (80問版) - 旧backupは触らない
cur = json.load(open(DATA, encoding="utf-8"))
with open(os.path.join(ROOT,"data","questions_backup_80問_2026-06-10_pre-71反映.json"),"w",encoding="utf-8") as f:
    json.dump(cur, f, ensure_ascii=False, indent=2)
print(f"[backup] 編集前 {len(cur)} 問 -> questions_backup_80問_2026-06-10_pre-71反映.json")

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
questions, problems = [], []
n_word = n_excel = 0

for r in range(2, ws.max_row+1):
    no   = ws.cell(r,1).value
    src  = ws.cell(r,2).value
    cat  = ws.cell(r,3).value
    q    = ws.cell(r,4).value
    ans  = ws.cell(r,5).value
    ansno= ws.cell(r,6).value
    opts = [ws.cell(r,7+i).value for i in range(5)]
    expl = ws.cell(r,12).value
    if (q is None or str(q).strip()=="") and (ans is None):
        continue
    q   = str(q).strip()
    ans = str(ans).strip() if ans is not None else ""
    opts= [str(o).strip() for o in opts if o is not None and str(o).strip()!=""]
    cat = str(cat).strip() if cat else "養豚一般"
    expl= str(expl).strip() if expl else ""
    is_word = bool(src) and ("Word" in str(src) or "本試験" in str(src) or "honshiken" in str(src))

    # --- 正答の整合 ---
    if ans not in opts:
        used = None
        try:
            i = int(ansno)
            if 1 <= i <= len(opts): used = opts[i-1]
        except (TypeError, ValueError):
            pass
        if used is not None and norm(used) != norm(ans):
            # 正答列と正答番号が食い違う → 報告（番号を採用しつつ要確認）
            best = max(opts, key=lambda o: difflib.SequenceMatcher(None,norm(o),norm(ans)).ratio())
            ratio = difflib.SequenceMatcher(None,norm(best),norm(ans)).ratio()
            if ratio >= 0.85:
                used = best  # 正答テキストにほぼ一致する選択肢を優先
            problems.append(("No.%s"%no,"正答列='%s' 正答番号=%s→'%s' 採用='%s'"%(ans,ansno,opts[int(ansno)-1] if str(ansno).isdigit() and 1<=int(ansno)<=len(opts) else "?",used)))
        elif used is None:
            best = max(opts, key=lambda o: difflib.SequenceMatcher(None,norm(o),norm(ans)).ratio())
            ratio= difflib.SequenceMatcher(None,norm(best),norm(ans)).ratio()
            used = best
            problems.append(("No.%s"%no,"正答'%s'が選択肢に無い→類似%.2f '%s'採用"%(ans,ratio,best)))
        ans = used

    source = ("honshiken-%s"%no) if is_word else ("setA-%s"%no)
    if is_word: n_word += 1
    else:       n_excel += 1
    questions.append({
        "question": q, "options": opts, "answer": ans,
        "explanation": expl, "category": cat, "source": source,
    })

print(f"[build] 合計 {len(questions)} 問  (Word/本試験 {n_word} + Excel/セットA {n_excel})")
from collections import Counter
for c,n in Counter(x["category"] for x in questions).most_common():
    print(f"   {c}: {n}")
bad=[i for i,q in enumerate(questions,1) if q["answer"] not in q["options"]]
print("[check] answer不整合:", bad if bad else "なし(全問OK)")
if problems:
    print("\n[!] 要確認:")
    for p in problems: print("   ", p)

with open(DATA,"w",encoding="utf-8") as f:
    json.dump(questions, f, ensure_ascii=False, indent=2)
print(f"\n[write] {DATA} ({len(questions)} 問)")
